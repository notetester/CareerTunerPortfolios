"""NCS xlsx를 세분류 단위 계층 JSONL로 정규화한다.

정본 CLI::

    python ncs_normalize.py --input <source.xlsx> --sheet all --output <out.jsonl>

기존 ``python ncs_normalize.py <sheet|all> <out.jsonl>`` 호출도 유지한다. 이 경우
원본 경로는 ``NCS_SOURCE_XLSX`` 환경변수에서 읽는다.
"""

from __future__ import annotations

import argparse
import json
import os
from collections import OrderedDict
from pathlib import Path
from typing import Any, Iterable

from ncs_contract import classification_codes, code_text


def norm_sheet(ws: Any, sub_acc: OrderedDict[str, dict[str, Any]]) -> None:
    """worksheet 행을 ``ncs_code -> record`` 누적기에 합친다."""
    rows: Iterable[tuple[Any, ...]] = ws.iter_rows(values_only=True)
    for row_number, row in enumerate(rows, start=1):
        if row_number == 1:
            continue
        if len(row) < 20 or row[6] is None:
            continue

        major_code = code_text(row[0], "major.code")
        middle_code = code_text(row[2], "middle.code")
        minor_code = code_text(row[4], "minor.code")
        sub_code = code_text(row[6], "sub.code")
        ncs_code, canonical_sub_code = classification_codes(
            major_code, middle_code, minor_code, sub_code
        )

        sub = sub_acc.get(ncs_code)
        if sub is None:
            sub = {
                "major": {"code": major_code, "name": row[1]},
                "middle": {"code": middle_code, "name": row[3]},
                "minor": {"code": minor_code, "name": row[5]},
                # 각 level의 code는 해당 level-local 값만 보관한다.
                "sub": {"code": canonical_sub_code, "name": row[7]},
                "units": OrderedDict(),
            }
            sub_acc[ncs_code] = sub

        unit_no = code_text(row[8], "unitNo") if row[8] is not None else ""
        unit = sub["units"].get(unit_no)
        if unit is None:
            unit = {
                "unitNo": unit_no,
                "unitName": row[9],
                "level": row[10],
                "elements": OrderedDict(),
            }
            sub["units"][unit_no] = unit

        element_no = code_text(row[11], "elementNo") if row[11] is not None else ""
        element = unit["elements"].get(element_no)
        if element is None:
            element = {
                "elementNo": element_no,
                "elementName": row[12],
                "level": row[13],
                "criteria": OrderedDict(),
                "지식": OrderedDict(),
                "기술": OrderedDict(),
                "태도": OrderedDict(),
            }
            unit["elements"][element_no] = element

        if row[15]:
            criterion_no = code_text(row[14], "performanceCriterionNo")
            element["criteria"].setdefault(criterion_no, row[15])

        ksa_kind = str(row[17] or "").strip()
        if ksa_kind in ("지식", "기술", "태도") and row[19]:
            ksa_no = code_text(row[18], "ksaNo")
            element[ksa_kind].setdefault(ksa_no, row[19])


def to_record(sub: dict[str, Any]) -> dict[str, Any]:
    units = []
    for unit in sub["units"].values():
        elements = []
        for element in unit["elements"].values():
            elements.append(
                {
                    "elementNo": element["elementNo"],
                    "elementName": element["elementName"],
                    "level": element["level"],
                    "performanceCriteria": list(element["criteria"].values()),
                    "knowledge": list(element["지식"].values()),
                    "skills": list(element["기술"].values()),
                    "attitudes": list(element["태도"].values()),
                }
            )
        units.append(
            {
                "unitNo": unit["unitNo"],
                "unitName": unit["unitName"],
                "level": unit["level"],
                "elements": elements,
            }
        )
    return {
        "major": sub["major"],
        "middle": sub["middle"],
        "minor": sub["minor"],
        "sub": sub["sub"],
        "units": units,
    }


def write_jsonl(sub_acc: OrderedDict[str, dict[str, Any]], output: Path) -> dict[str, int]:
    counts = {
        "classifications": len(sub_acc),
        "units": 0,
        "elements": 0,
        "criteria": 0,
        "knowledge": 0,
        "skills": 0,
        "attitudes": 0,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as stream:
        for sub in sub_acc.values():
            record = to_record(sub)
            for unit in record["units"]:
                counts["units"] += 1
                for element in unit["elements"]:
                    counts["elements"] += 1
                    counts["criteria"] += len(element["performanceCriteria"])
                    counts["knowledge"] += len(element["knowledge"])
                    counts["skills"] += len(element["skills"])
                    counts["attitudes"] += len(element["attitudes"])
            stream.write(json.dumps(record, ensure_ascii=False) + "\n")
    return counts


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("legacy_sheet", nargs="?", help=argparse.SUPPRESS)
    parser.add_argument("legacy_output", nargs="?", help=argparse.SUPPRESS)
    parser.add_argument("--input", dest="source", help="NCS 원본 xlsx 경로")
    parser.add_argument("--sheet", help="worksheet 이름 또는 all(기본)")
    parser.add_argument("--output", help="정규화 JSONL 출력 경로")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    source_value = args.source or os.getenv("NCS_SOURCE_XLSX")
    output_value = args.output or args.legacy_output
    sheet = args.sheet or args.legacy_sheet or "all"
    if not source_value:
        parser.error("--input 또는 NCS_SOURCE_XLSX가 필요합니다")
    if not output_value:
        parser.error("--output이 필요합니다")
    if args.sheet and args.legacy_sheet:
        parser.error("--sheet와 기존 positional sheet를 함께 지정할 수 없습니다")

    source = Path(source_value).expanduser()
    output = Path(output_value).expanduser()
    if not source.is_file():
        parser.error(f"원본 xlsx를 찾을 수 없습니다: {source}")

    try:
        import openpyxl
    except ImportError as exc:
        parser.error(f"openpyxl이 필요합니다: {exc}")

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        sheet_names = workbook.sheetnames if sheet.lower() == "all" else [sheet]
        missing = [name for name in sheet_names if name not in workbook.sheetnames]
        if missing:
            parser.error(f"worksheet를 찾을 수 없습니다: {', '.join(missing)}")
        sub_acc: OrderedDict[str, dict[str, Any]] = OrderedDict()
        for sheet_name in sheet_names:
            norm_sheet(workbook[sheet_name], sub_acc)
            print(f"sheet {sheet_name}: {len(sub_acc)} classifications", flush=True)
        counts = write_jsonl(sub_acc, output)
    finally:
        workbook.close()

    print(json.dumps({"output": str(output), **counts}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
